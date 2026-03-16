from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
from functools import wraps
import face_recognition
import cv2
import numpy as np
import os
import pandas as pd
from datetime import datetime
import base64

from database import init_db, add_employee, get_all_employees, delete_employee, mark_attendance, get_attendance_logs, get_db_connection, get_cursor, get_placeholder
from face_utils import encode_face_from_image, serialize_encoding, deserialize_encoding, match_face

app = Flask(__name__)
app.config['SECRET_KEY'] = 'fas-secret-2026'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB limit

# ─── Database Init ────────────────────────────────────────────────────────────
init_db()

# ─── In-memory face cache ─────────────────────────────────────────────────────
known_encodings = []
known_ids = []
id_to_name = {}

def reload_faces():
    global known_encodings, known_ids, id_to_name
    known_encodings, known_ids = [], []
    id_to_name = {}
    for emp in get_all_employees():
        eid = emp['employee_id']
        name = emp['name']
        known_ids.append(eid)
        id_to_name[eid] = name
        known_encodings.append(deserialize_encoding(emp['face_encoding']))

reload_faces()

# ─── Helpers ──────────────────────────────────────────────────────────────────
def b64_to_bytes(b64str):
    if ',' in b64str:
        b64str = b64str.split(',', 1)[1]
    return base64.b64decode(b64str)

# ─── Auth ─────────────────────────────────────────────────────────────────────
ADMIN_EMAIL = os.getenv('ADMIN_EMAIL', 'admin@keyafusion.com')
ADMIN_PASS = os.getenv('ADMIN_PASS', 'admin123')

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json or {}
        email = data.get('email')
        password = data.get('password')
        if email == ADMIN_EMAIL and password == ADMIN_PASS:
            session['logged_in'] = True
            return jsonify({"success": True, "message": "Welcome back!"})
        return jsonify({"success": False, "message": "Invalid credentials"}), 401
    
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/send_otp', methods=['POST'])
def send_otp():
    return jsonify({"success": False, "message": "Admin registration is disabled. Please sign in."}), 403

@app.route('/register', methods=['POST'])
def register():
    return jsonify({"success": False, "message": "Admin registration is disabled."}), 403

@app.route('/')
@login_required
def index():
    # If this service is meant only for users, redirect home to the user panel
    if os.getenv('APP_MODE') == 'USER':
        return redirect(url_for('user_panel'))
        
    employees = get_all_employees()
    all_logs   = get_attendance_logs()
    today      = datetime.now().strftime('%Y-%m-%d')
    today_logs = [l for l in all_logs if l['date'] == today]
    
    return render_template('index.html', 
                           total_employees=len(employees),
                           present_today=len(today_logs),
                           recent_logs=all_logs[:8])

@app.route('/employees')
@login_required
def view_employees():
    return render_template('employees.html', employees=get_all_employees())

@app.route('/add_employee')
@login_required
def add_employee_route():
    return render_template('add_employee.html')

@app.route('/attendance')
@login_required
def attendance():
    return render_template('attendance.html')

@app.route('/history')
@login_required
def history():
    return render_template('history.html', logs=get_attendance_logs())

@app.route('/user')
def user_panel():
    return render_template('user_attendance.html')

@app.route('/manifest.json')
def serve_manifest():
    return send_file('static/manifest.json')

@app.route('/sw.js')
def serve_sw():
    return send_file('static/sw.js')

# ─── API: Add employee with live face samples ──────────────────────────────────
@app.route('/api/add_employee', methods=['POST'])
def api_add_employee():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        print(f"Error parsing JSON: {e}")
        return jsonify(success=False, message=f'Invalid request: {e}')

    eid        = (data.get('employee_id') or '').strip()
    name       = (data.get('name') or '').strip()
    department = data.get('department', 'IT')
    phone      = data.get('phone', '')
    email      = data.get('email', '')
    frames_b64 = data.get('frames', [])   # list of base64 strings

    if not eid or not name:
        return jsonify(success=False, message='Employee ID and Name are required.')
    if not frames_b64:
        return jsonify(success=False, message='No face frames received.')

    # Extract encodings from each captured frame
    good_encodings = []
    for b64 in frames_b64:
        enc = encode_face_from_image(b64_to_bytes(b64))
        if enc is not None:
            good_encodings.append(enc)

    if len(good_encodings) < 3:
        return jsonify(success=False,
                       message=f'Only {len(good_encodings)} valid face frames detected out of {len(frames_b64)}. '
                               'Ensure good lighting and face is fully visible.')

    # Average the encodings → robust single representation
    avg_encoding = np.mean(good_encodings, axis=0)
    blob = serialize_encoding(avg_encoding)

    ok = add_employee(eid, name, department, phone, email, blob)
    if not ok:
        return jsonify(success=False, message='Employee ID already exists. Use a different ID.')

    reload_faces()
    return jsonify(success=True,
                   message=f'✅ {name} registered with {len(good_encodings)} face samples!')

# ─── API: Delete employee ──────────────────────────────────────────────────────
@app.route('/api/delete_employee/<eid>', methods=['POST'])
def api_delete_employee(eid):
    delete_employee(eid)
    reload_faces()
    return jsonify(success=True)


# ─── API: Recognise a single frame ────────────────────────────────────────────
@app.route('/api/recognise', methods=['POST'])
def api_recognise():
    data = request.get_json(force=True)
    img_b64 = data.get('frame', '')
    if not img_b64:
        return jsonify(success=False, faces=[])

    img_bytes = b64_to_bytes(img_b64)
    nparr = np.frombuffer(img_bytes, np.uint8)
    bgr = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if bgr is None:
        return jsonify(success=False, faces=[])

    # Shrink for speed
    small = cv2.resize(bgr, (0, 0), fx=0.25, fy=0.25)
    rgb   = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)

    locs  = face_recognition.face_locations(rgb, model='hog')
    encs  = face_recognition.face_encodings(rgb, locs)

    results = []
    for (top, right, bottom, left), enc in zip(locs, encs):
        mid = match_face(known_encodings, known_ids, enc)
        name = id_to_name.get(mid, 'Unknown') if mid else 'Unknown'
        results.append({
            'id':  mid if mid else 'Unknown',
            'name': name,
            'box': {'top': top*4, 'right': right*4, 'bottom': bottom*4, 'left': left*4}
        })

    return jsonify(success=True, faces=results)


# ─── API: Mark attendance ─────────────────────────────────────────────────────
@app.route('/api/mark', methods=['POST'])
def api_mark():
    data = request.get_json(force=True)
    eid  = (data.get('employee_id') or '').strip()
    if not eid:
        return jsonify(success=False, message='No employee_id')
    status = mark_attendance(eid)
    return jsonify(success=True, message=status)

# ─── API: Manual attendance override ──────────────────────────────────────────
@app.route('/api/manual_attendance', methods=['POST'])
def api_manual_attendance():
    data = request.get_json(force=True)
    eid = data.get('employee_id')
    status = data.get('status')
    date_val = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    if not eid or not status:
        return jsonify(success=False, message='Missing parameters')
        
    conn = get_db_connection()
    cursor = get_cursor(conn)
    p = get_placeholder()
    
    # Check if record exists
    cursor.execute(f"SELECT id FROM attendance WHERE employee_id = {p} AND date = {p}", (eid, date_val))
    record = cursor.fetchone()
    
    if status == 'Present':
        in_time = datetime.now().strftime('%H:%M:%S')
        out_time = ''
    elif status == 'Absent':
        in_time = 'Absent'
        out_time = 'Absent'
    else:
        # Leave statuses (Sick Leave, Paid Leave)
        in_time = status
        out_time = status

    rec_id = record['id'] if record else None
    # SQLite row object behaves like dict but accessing by 'id' might need dict-like access
    if record and type(record) is not dict:
        rec_id = record[0] if isinstance(record, tuple) else record['id']

    if record:
        cursor.execute(f"UPDATE attendance SET login_time = {p}, logout_time = {p} WHERE id = {p}", (in_time, out_time, rec_id))
    else:
        cursor.execute(f"INSERT INTO attendance (employee_id, date, login_time, logout_time) VALUES ({p}, {p}, {p}, {p})", (eid, date_val, in_time, out_time))
        
    conn.commit()
    conn.close()
    return jsonify(success=True)


# ─── Export Excel ─────────────────────────────────────────────────────────────
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

@app.route('/export_excel')
def export_excel():
    logs = get_attendance_logs()
    employees = get_all_employees()
    
    # Setup Date Range for Current Month (up to today)
    today = datetime.now()
    start_date = today.replace(day=1)
    date_range = pd.date_range(start=start_date, end=today)
    
    # 1. Prepare Base Employee DataFrame
    emp_list = [{'ID': e['employee_id'], 'Name': e['name'], 'Department': e['department']} for e in employees]
    df_emp = pd.DataFrame(emp_list)
    
    if df_emp.empty:
      # Fallback if no employees
      df_emp = pd.DataFrame(columns=['ID', 'Name', 'Department'])
    
    # 2. Process Logs into a Dictionary
    # Key: (employee_id, date_str), Value: Status string
    attendance_dict = {}
    for l in logs:
        try:
           log_date = datetime.strptime(l['date'], '%Y-%m-%d')
           if log_date.month == today.month and log_date.year == today.year:
               eid = l['employee_id']
               if l['login_time'] == 'Absent':
                   status = 'Absent'
               else:
                   in_t = l.get('login_time', '-')
                   out_t = l.get('logout_time', '-') or '-'
                   status = f"IN: {in_t}\nOUT: {out_t}"
               attendance_dict[(eid, log_date.strftime('%Y-%m-%d'))] = status
        except Exception:
           pass
           
    # 3. Build Date Columns
    for single_date in date_range:
        date_str = single_date.strftime('%Y-%m-%d')
        col_name = single_date.strftime('%d-%b')
        is_sunday = single_date.weekday() == 6
        
        status_list = []
        for idx, row in df_emp.iterrows():
            eid = row['ID']
            if is_sunday:
                status_list.append('Holiday')
            else:
                status = attendance_dict.get((eid, date_str), 'Absent')
                status_list.append(status)
                
        df_emp[col_name] = status_list

    # 4. Save via OpenPyXL and Apply Colors
    path = os.path.join(os.path.dirname(__file__), 'data', 'Monthly_Attendance.xlsx')
    
    writer = pd.ExcelWriter(path, engine='openpyxl')
    df_emp.to_excel(writer, sheet_name='Attendance', index=False)
    
    workbook = writer.book
    worksheet = writer.sheets['Attendance']
    
    # Define Fills
    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    holiday_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    sick_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # light yellow/orange
    paid_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid") # light purple
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    header_font = Font(color="FFFFFF", bold=True)
    
    # Format Headers
    for col_num, col_name in enumerate(df_emp.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        col_let = get_column_letter(col_num)
        if col_num <= 3:
           worksheet.column_dimensions[col_let].width = 20
        else:
           worksheet.column_dimensions[col_let].width = 18 # More room for IN/OUT
    
    # Format Cells based on text
    for r_idx in range(2, len(df_emp) + 2):
        worksheet.row_dimensions[r_idx].height = 42 # Much taller for better spacing
        for c_idx in range(4, len(df_emp.columns) + 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            val = str(cell.value) if cell.value else ""
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if "IN:" in val:
                cell.fill = present_fill
                cell.font = Font(color="065F46", bold=True, size=10) # Slightly larger font
            elif val == 'Absent':
                cell.fill = absent_fill
                cell.font = Font(color="991B1B", size=9)
            elif val == 'Holiday':
                cell.fill = holiday_fill
                cell.font = Font(color="3730A3", size=9)
            elif val == 'Sick Leave':
                cell.fill = sick_fill
                cell.font = Font(color="92400E", bold=True, size=9)
            elif val == 'Paid Leave':
                cell.fill = paid_fill
                cell.font = Font(color="6B21A8", bold=True, size=9)
                
    writer.close()
    
    return send_file(path, as_attachment=True, download_name=f'Attendance_Report_{today.strftime("%b_%Y")}.xlsx')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True, threaded=True)
